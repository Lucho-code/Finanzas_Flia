"""
Generación de reportes XLSX y envío de email.
Compartido entre bot.py y admin_panel.py.
"""
import io
import os
import smtplib
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TOT_FILL = PatternFill("solid", fgColor="1F4E79")
_TOT_FONT = Font(bold=True, color="FFFFFF")
_OVER_FILL = PatternFill("solid", fgColor="FFC7CE")   # categoría sobre presupuesto
_INGRESO_FILL = PatternFill("solid", fgColor="C6EFCE")
_GASTO_FILL = PatternFill("solid", fgColor="FCE4D6")


def _style_hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def build_xlsx(db, records: list, titulo: str, start_date: date, end_date: date) -> io.BytesIO:
    """
    Genera el XLSX de finanzas familiares:
      - Hoja "Resumen": ingresos, gastos y saldo por miembro, y por categoría vs presupuesto.
      - Una hoja por miembro con el detalle de movimientos.
    """
    wb = openpyxl.Workbook()

    # ── Hoja Resumen ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, titulo)
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center")

    ingresos_total = sum(r["amount"] for r in records if r["type"] == "ingreso")
    gastos_total   = sum(r["amount"] for r in records if r["type"] == "gasto")
    saldo_total    = ingresos_total - gastos_total

    ws.cell(3, 1, "Ingresos totales").font = Font(bold=True)
    ws.cell(3, 2, round(ingresos_total, 2)).number_format = "#,##0.00"
    ws.cell(4, 1, "Gastos totales").font = Font(bold=True)
    ws.cell(4, 2, round(gastos_total, 2)).number_format = "#,##0.00"
    ws.cell(5, 1, "Saldo").font = Font(bold=True)
    saldo_cell = ws.cell(5, 2, round(saldo_total, 2))
    saldo_cell.number_format = "#,##0.00"
    saldo_cell.fill = _INGRESO_FILL if saldo_total >= 0 else _OVER_FILL

    # Por miembro
    row = 7
    ws.cell(row, 1, "Por miembro").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(["Miembro", "Ingresos", "Gastos", "Saldo"], start=1):
        c = ws.cell(row, col, h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    members = sorted({r["member_name"] for r in records}) or [m["name"] for m in db.list_members()]
    for name in members:
        ing = sum(r["amount"] for r in records if r["member_name"] == name and r["type"] == "ingreso")
        gas = sum(r["amount"] for r in records if r["member_name"] == name and r["type"] == "gasto")
        ws.cell(row, 1, name)
        ws.cell(row, 2, round(ing, 2)).number_format = "#,##0.00"
        ws.cell(row, 3, round(gas, 2)).number_format = "#,##0.00"
        ws.cell(row, 4, round(ing - gas, 2)).number_format = "#,##0.00"
        row += 1

    # Por categoría (gastos) vs presupuesto
    row += 2
    ws.cell(row, 1, "Gastos por categoría vs. presupuesto").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(["Categoría", "Gastado", "Presupuesto", "Estado"], start=1):
        c = ws.cell(row, col, h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    row += 1
    for cat in db.get_spent_by_category(start_date, end_date, "gasto"):
        ws.cell(row, 1, cat["category_name"])
        ws.cell(row, 2, round(cat["total"], 2)).number_format = "#,##0.00"
        presupuesto = cat.get("budget_monthly")
        if presupuesto:
            ws.cell(row, 3, presupuesto).number_format = "#,##0.00"
            excedido = cat["total"] > presupuesto
            ws.cell(row, 4, "Excedido" if excedido else "OK")
            if excedido:
                for col in range(1, 5):
                    ws.cell(row, col).fill = _OVER_FILL
        row += 1

    for col, w in zip("ABCD", [24, 16, 16, 12]):
        ws.column_dimensions[col].width = w

    # ── Una hoja por miembro ─────────────────────────────────────────────
    for name in members:
        sname = name[:31]
        wsm = wb.create_sheet(title=sname)
        headers = ["Fecha", "Tipo", "Categoría", "Monto", "Descripción"]
        for col, h in enumerate(headers, start=1):
            wsm.cell(1, col, h)
        _style_hdr(wsm, 1, len(headers))

        cur_row = 2
        mem_recs = [r for r in records if r["member_name"] == name]
        for r in sorted(mem_recs, key=lambda x: x["date"]):
            wsm.cell(cur_row, 1, r["date"])
            wsm.cell(cur_row, 2, "Ingreso" if r["type"] == "ingreso" else "Gasto")
            wsm.cell(cur_row, 3, r.get("category_name") or "Sin categoría")
            amt_cell = wsm.cell(cur_row, 4, round(r["amount"], 2))
            amt_cell.number_format = "#,##0.00"
            wsm.cell(cur_row, 5, r.get("description") or "")
            fill = _INGRESO_FILL if r["type"] == "ingreso" else _GASTO_FILL
            for col in range(1, 6):
                wsm.cell(cur_row, col).fill = fill
            cur_row += 1

        total_row = cur_row
        ws_ing = sum(r["amount"] for r in mem_recs if r["type"] == "ingreso")
        ws_gas = sum(r["amount"] for r in mem_recs if r["type"] == "gasto")
        wsm.cell(total_row, 3, "TOTAL").font = _TOT_FONT
        wsm.cell(total_row, 3).fill = _TOT_FILL
        wsm.cell(total_row, 4, round(ws_ing - ws_gas, 2)).number_format = "#,##0.00"
        wsm.cell(total_row, 4).font = _TOT_FONT
        wsm.cell(total_row, 4).fill = _TOT_FILL

        for col, w in zip("ABCDE", [12, 10, 18, 12, 30]):
            wsm.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def send_email(buf: io.BytesIO, filename: str, subject: str) -> bool:
    email_from = os.getenv("EMAIL_FROM")
    email_pass = os.getenv("EMAIL_PASSWORD")
    email_to   = os.getenv("EMAIL_TO")
    if not all([email_from, email_pass, email_to]):
        return False
    msg = MIMEMultipart()
    msg["From"] = email_from
    msg["To"] = email_to
    msg["Subject"] = subject
    msg.attach(MIMEText(
        f"Adjunto el reporte de finanzas familiares: {filename}\n\n"
        "Enviado automáticamente por el sistema.",
        "plain", "utf-8",
    ))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(email_from, email_pass)
        s.sendmail(email_from, email_to.split(","), msg.as_string())
    return True
